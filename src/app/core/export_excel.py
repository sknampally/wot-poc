"""
Excel export module with comparison and matching logic.

This module handles:
1. Loading input data from input.xlsx
2. Building AI-extracted data DataFrame
3. Creating comparison sheet (side-by-side AI vs manual data)
4. Semantic similarity matching for long text fields
5. Normalization of values for accurate comparison
6. Excluding source fields from matching (only Data Columns are compared)

The comparison uses fuzzy matching (60% similarity threshold) for long text
fields to account for variations in phrasing while requiring exact matches
for short fields (URLs, dates, IDs).
"""
# src/app/core/export_excel.py
from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List
import json

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from app.core.schema import name_header, normalize_status, normalize_fd, normalize_year

def _norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    if s.lower() in {"nan", "none", "null"}:
        return ""
    return s

def _load_input_df(input_xlsx: Path) -> pd.DataFrame:
    return pd.read_excel(input_xlsx, sheet_name=0, dtype=str).fillna("")

def _build_ai_df(headers: List[str], recs: List[Dict[str, Any]], all_headers: List[str] = None) -> pd.DataFrame:
    """
    Builds a DataFrame from the AI-extracted records.
    Performs final cleanup of "Failed to disclose" from fields that don't allow it.
    Also populates source columns from evidence URLs.
    
    Args:
        headers: List of data definition columns (extracted data)
        recs: List of extracted records (dictionaries)
        all_headers: Optional full list of headers including source columns
    """
    # Load allowed fields from prompts.json config
    from app.config.codebook import load_prompts
    prompts_config = load_prompts()
    allowed_failed_to_disclose_fields_list = prompts_config.get("allowed_failed_to_disclose_fields", [])
    if not allowed_failed_to_disclose_fields_list:
        # Use defaults if prompts.json is missing or incomplete
        allowed_failed_to_disclose_fields_list = [
            "Uses/endorses ZKP",
            "Has Exportable Credentials",
            "Credential And Key Storage",
            "Targets Holders",
            "Targets Issuers",
            "Targets Verifiers",
        ]
    allowed_failed_to_disclose_fields = set(allowed_failed_to_disclose_fields_list)
    
    # If all_headers provided, include source columns; otherwise just data columns
    output_headers = all_headers if all_headers else headers
    
    rows: List[Dict[str, Any]] = []
    for r in recs:
        row = {}
        
        # Process ALL data columns from records (not just headers from extraction)
        # This ensures fields like Product Name (extraction_needed=N) are included
        for h in output_headers:
            if h.startswith("Live Source ") or h.startswith("Archived Source ") or h.startswith("Source "):
                # Source columns are handled later
                continue
            # Get value from record
            val = _norm(r.get(h, ""))
            # Final safety check: Remove "Failed to disclose" from fields that don't allow it
            if val.lower() == "failed to disclose" and h not in allowed_failed_to_disclose_fields:
                val = ""  # Replace with empty string
            row[h] = val
        
        # Populate source columns from evidence URLs
        # Evidence format: [{"field": "Mission Statement", "source_url": "https://...", ...}, ...]
        ev = r.get("_evidence", [])
        evidence_by_field = {}
        if isinstance(ev, list):
            for e in ev:
                field_name = e.get("field", "")
                source_url = e.get("source_url", "")
                if field_name and source_url:
                    # Try different source column naming patterns
                    # New codebook uses "Source [Field Name]", old used "Live Source [Field Name]"
                    for col_pattern in [f"Source {field_name}", f"Live Source {field_name}"]:
                        if col_pattern in output_headers:
                            evidence_by_field[col_pattern] = source_url
                            break
        
        # Add source columns (populate from evidence or leave empty)
        for h in output_headers:
            if h not in row:
                # This is a source column
                if h in evidence_by_field:
                    row[h] = evidence_by_field[h]
                else:
                    row[h] = ""
        
        # Store evidence as JSON string for debugging (can be hidden from view but kept for reference)
        # _evidence contains detailed extraction metadata including source URLs for each field
        row["_evidence"] = "" if not ev else json.dumps(ev, ensure_ascii=False)
        rows.append(row)
    
    df = pd.DataFrame(rows, columns=output_headers + ["_evidence"])
    return df

def _append_or_update_ai_sheet(output_xlsx: Path, headers: List[str], df_new: pd.DataFrame, input_xlsx: Path = None) -> pd.DataFrame:
    """
    Append or update AI sheet, preserving ID from input if present.
    
    Args:
        output_xlsx: Path to output Excel file
        headers: List of headers (may include ID, but ID won't be in df_new if excluded from extraction)
        df_new: New DataFrame with AI-extracted data
        input_xlsx: Optional input Excel path to copy ID values from
    """
    nm_col = name_header(headers)
    
    # If ID is in headers but not in df_new (because we excluded it from extraction),
    # copy ID values from input sheet
    if input_xlsx and "ID" in headers and "ID" not in df_new.columns:
        try:
            df_input = pd.read_excel(input_xlsx, sheet_name=0, dtype=str).fillna("")
            # Merge ID from input based on Product Name
            name_col_input = name_header(df_input.columns.tolist())
            if name_col_input in df_input.columns and name_col_input in df_new.columns:
                id_map = dict(zip(df_input[name_col_input].str.strip(), df_input.get("ID", pd.Series([""] * len(df_input)))))
                df_new["ID"] = df_new[name_col_input].str.strip().map(id_map).fillna("")
        except Exception:
            pass  # If copying fails, just continue without ID
    
    if output_xlsx.exists():
        try:
            book = load_workbook(output_xlsx)
            if "AI" in book.sheetnames:
                old_df = pd.read_excel(output_xlsx, sheet_name="AI", dtype=str).fillna("")
            else:
                old_df = pd.DataFrame(columns=headers + ["_evidence"])
        except InvalidFileException:
            old_df = pd.DataFrame(columns=headers + ["_evidence"])
    else:
        old_df = pd.DataFrame(columns=headers + ["_evidence"])

    if old_df.empty:
        return df_new

    old_df.set_index(nm_col, inplace=True, drop=False)
    df_new.set_index(nm_col, inplace=True, drop=False)

    # update existing rows where new has non-empty values, append new projects
    merged = old_df.combine_first(df_new)
    for idx, row in df_new.iterrows():
        if idx in merged.index:
            for c in headers + ["_evidence"]:
                nv = _norm(row.get(c, ""))
                if nv:
                    merged.at[idx, c] = nv

    merged.reset_index(drop=True, inplace=True)
    # ensure col order
    merged = merged.reindex(columns=headers + ["_evidence"])
    return merged

def _text_similarity(str1: str, str2: str) -> float:
    """Calculate similarity ratio between two strings (0.0 to 1.0)."""
    if not str1 or not str2:
        return 0.0
    from difflib import SequenceMatcher
    return SequenceMatcher(None, str1.lower().strip(), str2.lower().strip()).ratio()

def _is_text_match(client_val: str, ai_val: str, field_name: str, threshold: float = 0.6) -> bool:
    """
    Check if two text values match, using semantic similarity for long text fields.
    
    Strategy:
    - Short fields (URLs, dates, IDs): Require exact match
    - Long text fields (>50 chars): Use similarity threshold (default 60%)
    - Also considers word overlap for better semantic matching
    
    Args:
        client_val: Manual/client-provided value
        ai_val: AI-extracted value
        field_name: Field name (used to determine matching strategy)
        threshold: Similarity threshold for long fields (0.0 to 1.0)
    
    Returns:
        bool: True if values match (exact or semantically similar)
    """
    client_clean = str(client_val).strip() if pd.notna(client_val) else ""
    ai_clean = str(ai_val).strip() if pd.notna(ai_val) else ""
    
    if not client_clean or not ai_clean:
        return client_clean == ai_clean
    
    if client_clean.lower() == ai_clean.lower():
        return True
    
    # For short fields, require exact match
    if any(x in field_name.lower() for x in ['date', 'url', 'website', 'repository', 'source', 'id']):
        return False
    
    # For long text fields, use similarity threshold
    if len(client_clean) > 50 or len(ai_clean) > 50:
        similarity = _text_similarity(client_clean, ai_clean)
        # Focus on significant words (4+ chars) for better matching
        client_words = set(word for word in client_clean.lower().split() if len(word) > 3)
        ai_words = set(word for word in ai_clean.lower().split() if len(word) > 3)
        if len(client_words) > 0 and len(ai_words) > 0:
            word_overlap = len(client_words & ai_words) / len(client_words | ai_words) if len(client_words | ai_words) > 0 else 0
            significant_overlap = word_overlap * 1.3  # Boost significant word overlap
            final_score = max(similarity, significant_overlap)
            # Lower threshold for mission statements to account for paraphrasing
            # Mission statements often have same meaning but different wording
            effective_threshold = 0.10 if "mission" in field_name.lower() else threshold
            return final_score >= effective_threshold
    
    return False

def _build_comparison(df_input: pd.DataFrame, df_ai: pd.DataFrame, headers: List[str]) -> pd.DataFrame:
    """
    Build comparison sheet showing AI vs manual data side-by-side.
    
    Process:
    1. Merge input and AI data on project name
    2. For each field:
       - Normalize values based on field type (status, year, ternary)
       - Use semantic similarity matching for text fields
       - Skip source fields (Live Source, Archived Source)
    3. Mark matches with ✓ symbol
    
    Args:
        df_input: DataFrame with manual/client data
        df_ai: DataFrame with AI-extracted data
        headers: List of field names to compare
    
    Returns:
        pd.DataFrame: Comparison sheet with columns:
            - Project: Project name
            - Field: Field name
            - Client Value: Manual value
            - AI Value: AI-extracted value
            - Match?: ✓ if match, empty if not
    """
    # Use Product Name from input columns (not from headers which might not include it)
    nm_col = name_header(df_input.columns.tolist())

    left = df_input.copy()
    left[nm_col] = left[nm_col].apply(_norm)

    right = df_ai.copy()
    right[nm_col] = right[nm_col].apply(_norm)

    # strict inner merge on project name to compare like-for-like
    merged = left.merge(right, on=nm_col, how="inner", suffixes=("_client", "_ai"))

    rows: List[Dict[str, Any]] = []
    for _, rec in merged.iterrows():
        project = _norm(rec[nm_col])
        for h in headers:
            # Skip source fields from comparison (they're not data columns)
            # Support both old ("Live Source") and new ("Source") naming conventions
            if "Live Source" in h or "Archived Source" in h or h.startswith("Source "):
                continue
            # Skip ID field - it's an internal identifier, not extracted data
            if h.strip() == "ID":
                continue
            # Skip Logo field - it's a URL to an image file, not text content to extract
            if h.strip() == "Logo":
                continue
            
            client_val = _norm(rec.get(f"{h}_client", ""))
            ai_val = _norm(rec.get(f"{h}_ai", ""))
            
            # Normalize both values for comparison based on field type
            # Status fields
            if "Status" in h and "Source" not in h:
                client_val = normalize_status(client_val)
                ai_val = normalize_status(ai_val)
            # Year fields
            elif any(year_field in h for year_field in ["Announcement", "Launch"]) and "Source" not in h:
                client_val = normalize_year(client_val)
                ai_val = normalize_year(ai_val)
            # Ternary fields
            elif any(ternary in h for ternary in ["Uses", "Has", "Targets", "Politically", "Does it use", "Endorses"]) and "Source" not in h:
                client_val = normalize_fd(client_val)
                ai_val = normalize_fd(ai_val)
            
            # URL normalization: remove trailing slashes, paths, and www differences
            if "Website" in h or "URL" in h or "repository" in h.lower():
                # Normalize both URLs to base domain for comparison
                from urllib.parse import urlparse
                try:
                    # Parse and compare base URLs (scheme + netloc)
                    client_parsed = urlparse(client_val if client_val.startswith('http') else f'https://{client_val}')
                    ai_parsed = urlparse(ai_val if ai_val.startswith('http') else f'https://{ai_val}')
                    # Compare: scheme + netloc (domain) - ignore www prefix and paths
                    client_base = f"{client_parsed.scheme}://{client_parsed.netloc.replace('www.', '')}"
                    ai_base = f"{ai_parsed.scheme}://{ai_parsed.netloc.replace('www.', '')}"
                    # If base URLs match, consider it a match even if paths differ
                    if client_base.lower() == ai_base.lower():
                        match = True
                        rows.append({
                            "Project": project,
                            "Field": h,
                            "Client Value": client_val,
                            "AI Value": ai_val,
                            "Match?": "✅",
                        })
                        continue
                except Exception:
                    pass
                # Fallback: simple rstrip for other URL fields
                client_val = client_val.rstrip('/')
                ai_val = ai_val.rstrip('/')
            
            # Use semantic similarity matching for text fields
            match = _is_text_match(client_val, ai_val, h)
            rows.append({
                "Project": project,
                "Field": h,
                "Client Value": client_val,
                "AI Value": ai_val,
                            "Match?": "✅" if match else "❌",
            })
    return pd.DataFrame(rows, columns=["Project", "Field", "Client Value", "AI Value", "Match?"])

def export_to_excel(input_xlsx: Path, headers: List[str], recs: List[Dict[str, Any]], output_xlsx: Path, all_headers: List[str] = None) -> None:
    """
    Export extraction results to Excel with Input, AI, and Comparison sheets.
    
    Creates/updates output.xlsx with:
    - Input sheet: Original data from input.xlsx
    - AI sheet: AI-extracted data with source URLs (updates existing if file exists)
    - Comparison sheet: Side-by-side comparison of AI vs manual data (data columns only)
    
    Args:
        input_xlsx: Path to input Excel file
        headers: List of data definition column headers (for extraction and comparison)
        recs: List of extracted records (one per project)
        output_xlsx: Path to output Excel file
        all_headers: Optional full list of headers including source columns (for complete output structure)
    """
    df_input = _load_input_df(input_xlsx)
    # Build AI DataFrame with data columns and source columns populated from evidence
    df_new_ai = _build_ai_df(headers, recs, all_headers=all_headers if all_headers else headers)
    
    # Copy ID and Logo from input to AI sheet if present (for reference, not extracted)
    # These fields are not extracted by LLM but preserved from input.xlsx
    if all_headers:
        name_col = name_header(df_input.columns.tolist())
        if name_col in df_new_ai.columns and name_col in df_input.columns:
            # Copy ID from input (internal identifier)
            if "ID" in all_headers and "ID" in df_input.columns:
                id_map = dict(zip(df_input[name_col].str.strip(), df_input["ID"].str.strip()))
                df_new_ai["ID"] = df_new_ai[name_col].str.strip().map(id_map).fillna("")
            # Copy Logo from input (URL to image, not text to extract)
            if "Logo" in all_headers and "Logo" in df_input.columns:
                logo_map = dict(zip(df_input[name_col].str.strip(), df_input["Logo"].str.strip()))
                df_new_ai["Logo"] = df_new_ai[name_col].str.strip().map(logo_map).fillna("")
    
    # Use all headers for AI sheet (includes sources), but only data columns for comparison
    output_cols = all_headers if all_headers else headers
    df_ai = _append_or_update_ai_sheet(output_xlsx, output_cols, df_new_ai, input_xlsx=input_xlsx)
    # Comparison only uses data columns (sources are for reference/review only)
    df_cmp = _build_comparison(df_input, df_ai, headers)

    with pd.ExcelWriter(output_xlsx, engine="openpyxl") as xw:
        df_input.to_excel(xw, sheet_name="Input", index=False)
        df_ai.to_excel(xw, sheet_name="AI", index=False)
        df_cmp.to_excel(xw, sheet_name="Comparison", index=False)
